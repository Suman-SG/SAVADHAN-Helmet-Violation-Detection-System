"""
DATABASE MODULE
SQLite database for registered vehicles and violation records
"""

import sqlite3
import os
from datetime import datetime
import config

class TrafficDatabase:
    def __init__(self, db_path=None):
        self.db_path = db_path or config.DATABASE_PATH
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        self._init_tables()
        self._load_sample_data()
    
    def _init_tables(self):
        """Create all required tables"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Registered vehicles table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS registered_vehicles (
                plate_number TEXT PRIMARY KEY,
                owner_name TEXT NOT NULL,
                owner_email TEXT NOT NULL,
                phone TEXT,
                address TEXT,
                vehicle_model TEXT,
                registration_date TIMESTAMP,
                is_active BOOLEAN DEFAULT 1
            )
        ''')
        
        # Violations table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS violations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plate_number TEXT NOT NULL,
                violation_date TIMESTAMP NOT NULL,
                violation_type TEXT NOT NULL,
                evidence_path TEXT,
                fine_amount REAL,
                status TEXT DEFAULT 'PENDING',
                email_sent BOOLEAN DEFAULT 0,
                email_sent_time TIMESTAMP NULL,
                invoice_path TEXT,
                FOREIGN KEY (plate_number) REFERENCES registered_vehicles(plate_number)
            )
        ''')
        
        # Invoices table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                violation_id INTEGER,
                invoice_number TEXT UNIQUE,
                pdf_path TEXT,
                generated_date TIMESTAMP,
                paid BOOLEAN DEFAULT 0,
                payment_date TIMESTAMP,
                FOREIGN KEY (violation_id) REFERENCES violations(id)
            )
        ''')
        
        conn.commit()
        # Ensure migrations for older DBs: add email_sent_time if missing
        try:
            cursor.execute("PRAGMA table_info(violations)")
            cols = [r[1] for r in cursor.fetchall()]
            if 'email_sent_time' not in cols:
                cursor.execute("ALTER TABLE violations ADD COLUMN email_sent_time TIMESTAMP NULL")
                print("[Database] Migrated: added column 'email_sent_time' to violations")
                conn.commit()
        except Exception:
            pass

        conn.close()
        print("[Database] Tables initialized")
    
    def _load_sample_data(self):
        """Load sample vehicle data for testing"""
        sample_vehicles = [
            ("MH12AB1234", "Rahul Sharma", "rahul.sharma@example.com", "9876543210", "Mumbai", "Honda Activa"),
            ("DL01CA1234", "Priya Singh", "priya.singh@example.com", "9876543211", "Delhi", "Suzuki Access"),
            ("KA01AB5678", "Amit Kumar", "amit.kumar@example.com", "9876543212", "Bangalore", "TVS Jupiter"),
            ("TN22CD9012", "Sneha Reddy", "sneha.reddy@example.com", "9876543213", "Chennai", "Hero Splendor"),
            ("GJ05XY3456", "Vikram Patel", "vikram.patel@example.com", "9876543214", "Ahmedabad", "Bajaj Pulsar"),
            ("UP16AB7890", "Neha Gupta", "neha.gupta@example.com", "9876543215", "Lucknow", "Royal Enfield"),
        ]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        for vehicle in sample_vehicles:
            try:
                cursor.execute('''
                    INSERT OR IGNORE INTO registered_vehicles 
                    (plate_number, owner_name, owner_email, phone, address, vehicle_model, registration_date, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                ''', (vehicle[0], vehicle[1], vehicle[2], vehicle[3], vehicle[4], vehicle[5], datetime.now()))
            except Exception:
                pass
        
        conn.commit()
        conn.close()
        print("[Database] Sample vehicle data loaded")
    
    def get_vehicle(self, plate_number):
        """Get vehicle owner details by plate number"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT plate_number, owner_name, owner_email, phone, address, vehicle_model
            FROM registered_vehicles 
            WHERE plate_number = ? AND is_active = 1
        ''', (plate_number.upper(),))
        
        result = cursor.fetchone()
        conn.close()
        
        if result:
            return {
                "plate_number": result[0],
                "owner_name": result[1],
                "owner_email": result[2],
                "phone": result[3],
                "address": result[4],
                "vehicle_model": result[5]
            }
        return None
    
    def register_vehicle(self, plate_number, owner_name, owner_email, phone=None, address=None, vehicle_model=None):
        """Register a new vehicle"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT OR REPLACE INTO registered_vehicles 
                (plate_number, owner_name, owner_email, phone, address, vehicle_model, registration_date, is_active)
                VALUES (?, ?, ?, ?, ?, ?, ?, 1)
            ''', (plate_number.upper(), owner_name, owner_email, phone, address, vehicle_model, datetime.now()))
            conn.commit()
            return True
        except Exception as e:
            print(f"[Database] Error registering vehicle: {e}")
            return False
        finally:
            conn.close()
    
    def record_violation(self, plate_number, violation_type, evidence_path, fine_amount):
        """Record a new violation"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO violations (plate_number, violation_date, violation_type, evidence_path, fine_amount, status, email_sent)
            VALUES (?, ?, ?, ?, ?, 'PENDING', 0)
        ''', (plate_number.upper(), datetime.now(), violation_type, evidence_path, fine_amount))
        
        violation_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        print(f"[Database] Violation recorded: ID={violation_id}, Plate={plate_number}")
        return violation_id
    
    def has_recent_violation(self, plate_number, hours=24):
        """Check if plate has violation in last X hours (prevent duplicate fines)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT COUNT(*) FROM violations 
            WHERE plate_number = ? 
            AND violation_date > datetime('now', ?)
            AND status = 'PENDING'
        ''', (plate_number.upper(), f'-{hours} hours'))
        
        count = cursor.fetchone()[0]
        conn.close()
        return count > 0
    
    def mark_email_sent(self, violation_id):
        """Mark that email was sent for a violation"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('UPDATE violations SET email_sent = 1, email_sent_time = ? WHERE id = ?', (datetime.now(), violation_id))
        conn.commit()
        conn.close()

    def get_last_email_sent_for_plate(self, plate_number):
        """Return datetime of last email sent for a plate, or None"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT email_sent_time FROM violations
            WHERE plate_number = ? AND email_sent_time IS NOT NULL
            ORDER BY email_sent_time DESC LIMIT 1
        ''', (plate_number.upper(),))
        row = cursor.fetchone()
        conn.close()
        if not row or row[0] is None:
            return None
        try:
            # SQLite returns ISO string; parse to datetime
            return datetime.fromisoformat(row[0]) if isinstance(row[0], str) else datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
        except Exception:
            try:
                return datetime.fromisoformat(row[0])
            except Exception:
                return None
    
    def mark_invoice_generated(self, violation_id, invoice_number, pdf_path):
        """Record invoice generation"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO invoices (violation_id, invoice_number, pdf_path, generated_date, paid)
            VALUES (?, ?, ?, ?, 0)
        ''', (violation_id, invoice_number, pdf_path, datetime.now()))
        conn.commit()
        conn.close()
    
    def get_all_vehicles(self):
        """Get all registered vehicles (for Excel export)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT plate_number, owner_name, owner_email, phone, address, vehicle_model FROM registered_vehicles WHERE is_active = 1')
        results = cursor.fetchall()
        conn.close()
        
        return [{"plate_number": r[0], "owner_name": r[1], "owner_email": r[2], 
                 "phone": r[3], "address": r[4], "vehicle_model": r[5]} for r in results]
    
    def get_all_violations(self):
        """Get all violations (for report)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT v.id, v.plate_number, v.violation_date, v.violation_type, 
                   v.fine_amount, v.status, v.email_sent, r.owner_name, r.owner_email
            FROM violations v
            LEFT JOIN registered_vehicles r ON v.plate_number = r.plate_number
            ORDER BY v.violation_date DESC
        ''')
        results = cursor.fetchall()
        conn.close()
        
        return [{"id": r[0], "plate": r[1], "date": r[2], "type": r[3], 
                 "fine": r[4], "status": r[5], "email_sent": r[6],
                 "owner": r[7], "email": r[8]} for r in results]
    
    def export_to_excel(self, output_path="vehicle_report.xlsx"):
        """Export all data to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Sheet 1: Vehicles
            ws1 = wb.active
            ws1.title = "Registered Vehicles"
            headers1 = ["Plate Number", "Owner Name", "Email", "Phone", "Address", "Vehicle Model"]
            ws1.append(headers1)
            
            for v in self.get_all_vehicles():
                ws1.append([v["plate_number"], v["owner_name"], v["owner_email"], 
                           v.get("phone", ""), v.get("address", ""), v.get("vehicle_model", "")])
            
            # Sheet 2: Violations
            ws2 = wb.create_sheet("Violations")
            headers2 = ["ID", "Plate Number", "Date", "Type", "Fine", "Status", "Email Sent", "Owner"]
            ws2.append(headers2)
            
            for v in self.get_all_violations():
                ws2.append([v["id"], v["plate"], v["date"], v["type"], 
                           v["fine"], v["status"], "Yes" if v["email_sent"] else "No", v.get("owner", "Unknown")])
            
            # Style headers
            for ws in [ws1, ws2]:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            wb.save(output_path)
            print(f"[Database] Exported to {output_path}")
            return output_path
        except ImportError:
            print("[Database] openpyxl not installed. Run: pip install openpyxl")
            return None


# Quick test
if __name__ == "__main__":
    db = TrafficDatabase()
    print("\n=== Registered Vehicles ===")
    for v in db.get_all_vehicles():
        print(f"  {v['plate_number']}: {v['owner_name']} ({v['owner_email']})")
    
    print("\n=== Test Lookup ===")
    vehicle = db.get_vehicle("MH12AB1234")
    if vehicle:
        print(f"  Found: {vehicle}")